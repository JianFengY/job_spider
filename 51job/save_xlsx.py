"""
Created on 2018/8/13
@Author: Jeff Yang
"""

import openpyxl
import json

wb = openpyxl.Workbook()

ws = wb.active
ws['A1'] = '职位名称'
ws['B1'] = '职位URL'
ws['C1'] = '工作地点'
ws['D1'] = '薪资'
ws['E1'] = '发布日期'
ws['F1'] = '公司名称'
ws['G1'] = '公司性质'
ws['H1'] = '公司规模'
ws['I1'] = '公司业务范围'
ws['J1'] = '公司简介'

i = 1
with open('job_info.json', 'r', encoding='utf-8') as f:
    while True:
        line = f.readline()
        data = []
        if line:
            dict = json.loads(line)
            data.append(dict['job_title'])
            data.append(dict['job_url'])
            data.append(dict['work_place'])
            data.append(dict['salary'])
            data.append(dict['publish_time'])
            data.append(dict['company']['name'])
            data.append(dict['company']['nature'])
            data.append(dict['company']['scale'])
            data.append(dict['company']['business'])
            data.append(dict['company']['profile'])
            ws.append(data)
            print('第', i, '条……')
            i += 1
        else:
            break

wb.save('result.xlsx')