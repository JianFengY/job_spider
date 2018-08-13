"""
Created on 2018/8/13
@Author: Jeff Yang
"""

import openpyxl
import json

wb = openpyxl.load_workbook('result.xlsx')
ws = wb.create_sheet('拉勾', index=1)

ws['A1'] = '职位名称'
ws['B1'] = '薪资'
ws['C1'] = '城市'
ws['D1'] = '发布时间'
ws['E1'] = '公司名称'
ws['F1'] = '公司性质'
ws['G1'] = '公司规模'
ws['H1'] = '公司业务范围'

i = 1
with open('job_info.json', 'r', encoding='utf-8') as f:
    while True:
        line = f.readline()
        data = []
        if line:
            dict = json.loads(line)
            data.append(dict['position_name'])
            data.append(dict['salary'])
            data.append(dict['city'])
            data.append(dict['create_time'])
            data.append(dict['company_name'])
            data.append(dict['finance_stage'])
            data.append(dict['company_size'])
            data.append(dict['industry_field'])
            ws.append(data)
            print('第', i, '条……')
            i += 1
        else:
            break

wb.save('result.xlsx')